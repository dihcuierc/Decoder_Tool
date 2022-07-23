import os
import tkinter
import openpyxl.utils.exceptions
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
import sys
from tkinter import filedialog, Tk, Label, Entry, Button
from tkPDFViewer import tkPDFViewer as Pdf
from datetime import datetime
from re import search, findall

# dictionary with the RoB codes and its associated headers in each RoB code.
# Eg: In Vin Discrepancy Abnormality Detection, it contains 0501, 0502, 0507, 5001.
# Inside 0501, title is Timestamp, and it contains Length, Trip Counter, Time counter and Master Sync Info.
# The number is the number of cells it requires. Titles are all 2.
Rob_codes = {'0x050x01': ['Timestamp', 'Length 1', 'Trip Counter 2', 'Time Counter 4', 'Master Sync Info 1'],
             '0x050x02': ['Accumulated Distance', 'Length 1', 'Unit 1', 'Cumulative Mileage 3'],
             '0x050x07': ['Time', 'Length 1', 'Year 1', 'Month 1', 'Day 1', 'Hour 1', 'Minute 1', 'Second 1'],
             '0x500x01': ['+B voltage', 'Length 1', 'Voltage Data 2'],
             '0x050x03': ['Wakeup Fail (DCM fail)', 'Length 1'],
             '0x050x04': ['Sleep Fail (DCM fail)', 'Length 1'],
             '0x050x24': ['Wakeup Fail (Global-ECU)', 'Length 1'],
             '0x050x25': ['Sleep Fail (Global-ECU)', 'Length 1'],
             '0x050x26': ['ECU Wakeup Fail', 'Length 1'],
             '0x050x27': ['ECU Sleep Fail', 'Length 1'],
             '0x050x08': ['IP header', 'Length 1', 'IP Header 20'],
             '0x050x09': ['Transport layer header', 'Length 1', 'Transport Layer Header 20'],
             '0x050x0A': ['Domain Name', 'Length 1', 'Domain Name 27'],
             '0x050x0B': ['TLS error factor ID', 'Length 1', 'TLS Error Factor 1'],
             '0x050x0C': ['Tool authentication error factor', 'Length 1', 'Error 1'],
             '0x050x12': ['Source IP address', 'Length 1', 'IP Address 4'],
             '0x050x13': ['Destination IP address', 'Length 1', 'IP Address 4'],
             }

# the RoB codes for Wakeup Fail(DCM fail) (0503), Sleep Fail(DCM fail) (0504), Wakeup Fail(Global-ECU) (0524),
# Sleep Fail (Global-ECU) (0525), ECU Wakeup Fail(0526), ECU Sleep Fail (0527) has to be decoded even more than the
# other robs. Add on if there are more RoB that have the same format as the following RoBs
wakeups_rob_code = ['0x050x03', '0x050x04', '0x050x24', '0x050x25', '0x050x26', '0x050x27']

# the subcomponents in each of the wakeup binary decoding
# The 0 are fillers because the Rob is 20 bytes long but information is only 8 bytes long
wakeup_fail_etc = (
    ('Remote warning 1', 'Remote DTC FFD 1', 'Remote monitoring 2', 'Remote DDR 1',
     'Center request command acquisition (SMS) 3', 'Alarm notification 2', 'Remote immobilizer 2',
     'Stolen vehicle tracking/SVT 2'),
    ('Remote warning 1', 'Remote DTC FFD 1', 'Remote monitoring 2', 'Remote DDR 1',
     'Center request command acquisition (SMS) 3', 'Alarm notification 2', 'Remote immobilizer 2',
     'Stolen vehicle tracking/SVT 2'),
    ('Maintenance assistant 2', 'Remote diagnostic assistant 2', 'Operation forgotten remote confirmation 3',
     'Reject under repair 2', 'Contrast function 1', 'Last file 1', 'Remote control (hazard)*1 2',
     'Remote control (door)*1 2'),
    ('Maintenance assistant 2', 'Remote diagnostic assistant 2', 'Operation forgotten remote confirmation 3',
     'Reject under repair 2', 'Contrast function 1', 'Last file 1', 'Remote control (hazard)*1 2',
     'Remote control (door)*1 2'),
    ('Remote control (PW)*1 2', 'Remote control (remote start) 2', 'Operation monitoring (UBI) 2', 'eConnect 1',
     'Emergency call 1', 'CAN VP(via CAN) 1', 'VP+(via Ethernet 1', 'H/U communication 2'),
    ('Remote control (PW)*1 2', 'Remote control (remote start) 2', 'Operation monitoring (UBI) 2', 'eConnect 1',
     'Emergency call 1', 'CAN VP(via CAN) 1', 'VP+(via Ethernet 1', 'H/U communication 2'),
    ('Simple proxy function 2', 'Diagnostic Detection 2', 'Diagnostic Communication 2', 'Program Update 1',
     'NEV service 1', 'Collision management 2', 'Tow away alert function 2', 'USB Connection 1',
     'Unregistered Feature 2'),
    ('Simple proxy function 2', 'Diagnostic Detection 2', 'Diagnostic Communication 2', 'Program Update 1',
     'NEV service 1', 'Collision management 2', 'Tow away alert function 2', 'USB Connection 1',
     'Unregistered Feature 2'),
    ('0', '0', '0', '0', '0', '0', '0', '0'), ('0', '0', '0', '0', '0', '0', '0', '0'),
    ('0', '0', '0', '0', '0', '0', '0', '0'), ('0', '0', '0', '0', '0', '0', '0', '0'),
    ('0', '0', '0', '0', '0', '0', '0', '0'), ('0', '0', '0', '0', '0', '0', '0', '0'),
    ('0', '0', '0', '0', '0', '0', '0', '0'), ('0', '0', '0', '0', '0', '0', '0', '0'),
    ('0', '0', '0', '0', '0', '0', '0', '0'), ('0', '0', '0', '0', '0', '0', '0', '0'),
    ('0', '0', '0', '0', '0', '0', '0', '0'), ('0', '0', '0', '0', '0', '0', '0', '0'),
)

Responses = {'0': 'Normal', '1': 'Normal', 'ef': 'Not Transmitted', "ff": "Undefined command", "fe": "No response",
             "fd": "Unused", "fc": "Negative"}

trigger_dict = {'w': 'Warning', 'i': 'Ignition', 'c': 'Center', 'a': 'Unable to detect'}


def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)


# tkinter widget
# nothing much to add on here
def window_explorer():
    global save_address, open_address

    def browse_files():
        global filename
        excels = r"*.xlsx *.xlsm *.xltx *.xltm, *.txt"
        open_address1 = filedialog.askopenfilename(title="Select a File", initialdir=os.path.dirname(sys.executable),
                                                   filetypes=(("Accepted files", excels), ("All files", "*.*")))
        entry_open.delete(0, 'end')
        entry_open.insert(0, open_address1)
        filename = os.path.split(open_address1)[1].split('.')[0]

    def save_files():
        save_address1 = filedialog.askdirectory(initialdir=os.path.dirname(sys.executable), title="Select a File")
        entry_save.delete(0, 'end')
        entry_save.insert(0, save_address1)

    def convert():
        global save_address, open_address, passed, address_errors, text_errors, xlsx, date_time_str, choice
        open_address = entry_open.get()
        save_address = entry_save.get()
        if open_address:
            try:
                with open(open_address, 'r') as f:
                    f.readlines()
                    f.close()
                    choice = 1
            except (UnicodeError, FileNotFoundError):
                try:
                    openpyxl.load_workbook(filename=open_address, read_only=True).close()
                    choice = 2
                except Exception:
                    text_error(f'Error: {text_errors[0]}')
                    return
        if save_address:
            # checking that the address is accessible by the exe file or is it blocked by some admin if not user will
            # be prompted to input another destination
            try:
                if choice == 1:
                    xlsx.save(f'{save_address}/RoB-({filename})-{date_time_str}.xlsx')
                elif choice == 2:
                    xlsx.save(f'{save_address}/DTC-({filename})-{date_time_str}.xlsx')
            except Exception:
                address_error(f'Error: {address_errors[1]}')
                return
        if save_address == '' and open_address == '':
            address_error(f'Error: {text_errors[0][:-1]} and\n{address_errors[1][6:]}')
            return
        if save_address == '':
            address_error(f'Error: {address_errors[0]}')
            return
        if open_address == '':
            text_error(f'Error: {text_errors[0]}')
            return
        passed = 1
        window.destroy()

    def info():
        help_window = tkinter.Toplevel()
        help_window_width = 980
        help_window_height = 700
        image_dir = resource_path("programming.ico")
        help_window.iconbitmap(image_dir)
        help_width = window.winfo_screenwidth()
        help_height = window.winfo_screenheight()
        help_x = int(window.winfo_x() - help_width/10)
        help_y = int(window.winfo_y() - help_height/10)
        help_window.geometry(f'{help_window_width}x{help_window_height}+{help_x}+{help_y}')
        label = Label(help_window, text='If nothing appears, refer to User_manual.pdf', font=('Arial', 14))
        label.pack()
        v1 = Pdf.ShowPdf()
        v1.img_object_li.clear()
        v2 = v1.pdf_view(help_window, pdf_location=resource_path("User_manual.pdf"), width=980, height=700)
        v2.pack()
        help_window.mainloop()

    # text box error will pop up when user clicks convert if user does not put in an appropriate address to save the
    # Excel file
    def address_error(text):
        global passed

        def des():
            address_window.destroy()

        passed = 0
        address_window = Tk()
        address_window.title('Decoder')
        image_dir = resource_path("programming.ico")
        address_window.iconbitmap(image_dir)
        address_window_width = 270
        address_window_height = 110
        address_x = int(window.winfo_x() + address_window_width/5*4)
        address_y = int(window.winfo_y() + address_window_height)
        address_window.geometry(f'{address_window_width}x{address_window_height}+{address_x}+{address_y}')
        warning = Label(address_window, text=text, font=('Arial', 11), bg='white', width=30, height=4)
        warning.place(x=0, y=0)
        ok_button = Button(address_window, text="OK", command=des, relief='raised', bd=3, font=('Arial', 11), width=10)
        ok_button.place(x=85, y=75)
        address_window.mainloop()

    # another text box error window if user does not select a text file
    def text_error(text):
        global passed

        def des():
            text_window.destroy()

        passed = 0
        text_window = Tk()
        text_window.title('Decoder')
        image_dir = resource_path("programming.ico")
        text_window.iconbitmap(image_dir)
        text_window_width = 240
        text_window_height = 100
        text_x = int(window.winfo_x() + text_window_width/10*9)
        text_y = int(window.winfo_y() + text_window_height*1.2)
        text_window.geometry(f'{text_window_width}x{text_window_height}+{text_x}+{text_y}')
        warning = Label(text_window, text=text, font=('Arial', 11), bg='white', width=26,
                        height=3)
        warning.place(x=0, y=0)
        ok_button = Button(text_window, text="OK", command=des, relief='raised', bd=3, font=('Arial', 11), width=10)
        ok_button.place(x=70, y=61)
        text_window.mainloop()

    # main window
    window = Tk()
    window.title('Decoder')
    window_width = 700
    window_height = 320
    image_dir = resource_path("programming.ico")
    window.iconbitmap(image_dir)
    window.update()
    screen_x = window.winfo_screenwidth()
    screen_y = window.winfo_screenheight()
    window_x = int(screen_x / 2 - window_width / 2)
    window_y = int(screen_y / 2 - window_height / 2 - 50)
    window.geometry(f'{window_width}x{window_height}+{window_x}+{window_y}')

    title_label = Label(window, text="Decoder", font=('Arial', 30))
    title_label.place(x=260, y=25)
    open_label = Label(window, text="Input file location:", font=('Arial', 11))
    open_label.place(x=70, y=95)
    save_label = Label(window, text="Save file to:", font=('Arial', 11))
    save_label.place(x=70, y=175)

    entry_open = Entry(window, width=57, borderwidth=2, relief='groove', bd=3, font=('Arial', 11))
    entry_open.delete(0, 'end')
    entry_open.insert(0, open_address)
    button_browse = Button(window, text="Browse Files", command=browse_files, relief='raised', bd=3, font=('Arial', 11))
    entry_open.place(x=70, y=120, height=30)
    button_browse.place(x=550, y=117)

    entry_save = Entry(window, width=57, borderwidth=2, relief='groove', bd=3, font=('Arial', 11))
    button_save = Button(window, text="Browse Files", command=save_files, relief='raised', bd=3, font=('Arial', 11))
    entry_save.delete(0, 'end')
    entry_save.insert(0, save_address)
    entry_save.place(x=70, y=200, heigh=30)
    button_save.place(x=550, y=197)

    button_upload = Button(window, text="Convert", command=convert, relief='raised', height=2, width=50, bd=2,
                           bg='light blue', font=('Arial', 11))
    button_upload.place(x=70, y=250)
    button_help = Button(window, text="Help", command=info, relief='raised', height=2, width=10, bd=2,
                         bg='light blue', font=('Arial', 11))
    button_help.place(x=550, y=250)
    window.mainloop()


def dtc_in_rob_design(rob, dtcs):
    global total
    column = 1
    title = rob.cell(row=total, column=column)
    title.value = 'DTC'
    title.font = Font(bold=True)
    design(rob, title, start_row=total, end_row=total, start_column=column, end_column=column + 1, constant=2,
           style='thick')
    total += 1
    if len(dtcs['DTC']):
        for keys in dtcs.keys():
            column += 1
            header = rob.cell(row=total, column=column)
            header.value = keys
            header.font = Font(bold=True)
            if keys == 'DTC' or keys == 'DTC Statusbyte Confirmed':
                column += 1
        total += 1
        column = 2
        for i in range(len(dtcs['DTC'])):
            for keys in dtcs:
                rob.cell(row=total, column=column).value = dtcs[keys][i]
                column += 1
                if keys == 'DTC' or keys == 'DTC Statusbyte Confirmed':
                    column += 1
            column = 2
            total += 1

        for j in range(len(dtcs['DTC']) + 1):
            for i in range(11):
                if i == 0 or i == 3:
                    design(rob, rob.cell(column=2 + i, row=2 + j), start_row=2 + j, end_row=2 + j, start_column=2 + i,
                           end_column=3 + i, constant=2)
                    continue
                if i != 1 or i != 4:
                    design(rob, rob.cell(row=2 + j, column=2 + i), start_row=2 + j, start_column=2 + i,
                           end_column=2 + i, end_row=2 + j, constant=1)
    else:
        fail = rob.cell(row=total, column=column + 1)
        fail.value = 'No DTCs detected'
        design(rob, fail, start_row=total, end_row=total, start_column=column + 1, end_column=column + 3, constant=3)
        fail.font = Font(bold=True)
        total += 1


# reading the text file and saving it in a list 'lines'.
# positive_name_list: Names of RoBs with positive response, positive_occurrence: the line number that the RoB is at.
def text_file():
    global passed, open_address, save_address, address_errors, text_errors
    df = open(open_address, 'r')
    lines = df.readlines()
    df.close()
    positive_occurrences = []
    positive_name_list = []
    negative_name_list = []
    dtcs = {'DTC': [], 'DTC Statusbyte': [], 'DTC Statusbyte Confirmed': [],
            'Year': [], 'Month': [], 'Day': [], 'Hour': [], 'Minute': [], 'Seconds': []}
    # determining where the Rob is by looking for 'reportRoBSnapshotRecordByRecordNumber' and 'RoBCodeMaskRecord'
    # after that I look for positive/negative response then add it to the corresponding list.
    for index, line in enumerate(lines):
        lines[index] = line.strip()
        if 'DTCSnapshotRecordByDTCNumber' in line:
            z = 0
            length = 0
            while not search('\+B voltage', lines[z + index]):
                line_list = lines[z + index].split('\t')
                check_word = line_list[1].strip(' ')
                if line_list[0].strip(' ') == 'Tx' and check_word == 'DTC':
                    dtcs['DTC'].append(line_list[2].strip(' '))
                elif check_word == 'DTC Statusbyte':
                    dtcs['DTC Statusbyte'].append(line_list[2].strip(' '))
                elif check_word == 'DTC Statusbyte.Confirmed DTC':
                    dtcs['DTC Statusbyte Confirmed'].append(line_list[2].strip(' '))
                elif check_word == 'Year':
                    dtcs['Year'].append(line_list[2].strip(' '))
                elif check_word == 'Month':
                    dtcs['Month'].append(line_list[2].strip(' '))
                elif check_word == 'Day':
                    dtcs['Day'].append(line_list[2].strip(' '))
                elif check_word == 'Hour':
                    dtcs['Hour'].append(line_list[2].strip(' '))
                elif check_word == 'Minute':
                    dtcs['Minute'].append(line_list[2].strip(' '))
                elif check_word == 'Seconds':
                    dtcs['Seconds'].append(line_list[2].strip(' '))
                z += 1
            for key in dtcs.keys():
                try:
                    _ = dtcs[key][length]
                except IndexError:
                    dtcs[key].append('Unable to detect')
            length += 1
        if 'reportRoBSnapshotRecordByRecordNumber' in line:
            for i in range(0, 20):
                if 'RoBCodeMaskRecord' in lines[index + i]:
                    line_number = index + i
                    for j in range(0, 20):
                        if 'positive' in lines[index + i + j].lower():
                            positive_occurrences.append(line_number)
                            # just removing all the empty spaces if not it will be '    hi' like this
                            positive_name_list.append(' '.join(lines[index + i].split('\t')[-1].split()))
                            break
                        # negative_name_list: tuple containing(Name of RoB with negative response, Reason for failure)
                        if 'negative' in lines[index + i + j].lower() and 'response code' in lines[index + i + j + 1].lower():
                            negative_name_list.append((' '.join(lines[index + i].split('\t')[-1].split()),
                                                       ' '.join(lines[index + i + j + 1].split('\t')[-1].split())))
                            break
                    break
    return dtcs, positive_name_list, negative_name_list, lines, positive_occurrences


# default design
# I made a function for this because I realised that this code is used like 8 times
def design(rob, label, start_row, start_column, end_row, end_column, constant, style='thin'):
    rob.merge_cells(start_row=start_row, start_column=start_column, end_row=end_row,
                    end_column=end_column)
    label.alignment = Alignment(horizontal='center')
    for i in range(0, constant):
        rob.cell(row=start_row, column=start_column + i).border = Border(top=Side(border_style=style),
                                                                         right=Side(border_style=style),
                                                                         bottom=Side(border_style=style),
                                                                         left=Side(border_style=style))


# function to create each subheader e.g. 0501: Timestamp, Length, Trip Counter, Time Counter, Master Sync Info
def create_header(name, rob, row):
    column = 2
    table_title = rob.cell(row=row, column=column)
    table_title.value = Rob_codes[name][0]

    table_title.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type="solid")
    design(rob=rob, label=table_title, start_column=column, start_row=row, end_column=column + 1, end_row=row,
           constant=2)
    table_title.font = Font(bold=True)

    # creating headers for e.g. Length, Trip Counter, Time Counter, Master Sync Info
    column += 2
    for subheadings in Rob_codes[name][1:]:
        length = 0
        index = 0
        # determining the number of cells needed for the headers
        for position, letter in enumerate(subheadings):
            if letter.isdigit():
                length = length * 10 + int(letter)
                if index == 0:
                    index = position
        subheading_title = rob.cell(row=row, column=column)
        subheading_title.value = subheadings[:index]
        design(rob=rob, label=subheading_title, start_row=row, start_column=column,
               end_row=row, end_column=column + length - 1, constant=length)
        subheading_title.font = Font(bold=True)
        column += length


def create_numbers(rob, data, name):
    # number row is the column with 0x00, 0x00, 0x00, 0x01, 0x01, 0x01. Number row has to increment by 3 each time.
    # row is the current row that the table is being created
    # j is column
    # length limit is the number of bytes in the current table row eg: length limit of 0501
    # is 7 + 3(2 from RoB code and 1 from length)
    # count is the number of subheaders e.g. Trip counter is 1, time counter is 1, master sync info is 1
    # rob_string is e.g. 0x050x01. It reads the 1st 2 data to determine the code.
    # binary count, merge_list, merge_count, wakeup is related to wakeup fail decoding from binary to words.
    # the decoding is done in this function and the creating of headers too. it is not created in the header function
    # as this is an exception
    # instances is number to display 'Record #1' etc
    global total
    row = total + 3
    while data[0] != '    0x05':
        data = data[1:]
    j = 2
    count = 0
    length_limit = 0
    binary_count = 0
    merge_list = []
    merge_count = 0
    wakeup = ''
    rob_string = ''
    instances = 1
    instance_label = rob.cell(row=row - 2, column=2)
    instance_label.value = f'Record #{instances}'
    design(rob=rob, label=instance_label, start_row=row - 2, start_column=2,
           end_row=row - 2, end_column=11, constant=10, style='thick')
    instance_label.fill = PatternFill(start_color='FF8A09', end_color='FF8A09', fill_type="solid")
    instance_label.font = Font(bold=True)
    while data[0] is not None:
        if count == 0:
            rob_string = data[0].strip(' ') + data[1].strip(' ')
            if rob_string in Rob_codes.keys():
                create_header(rob_string, rob, row)
            else:
                minus = 0
                while data[0] != '    0x05' or data[1] != '    0x01':
                    data = data[1:]
                    minus += 1
                if minus != 3:
                    rob.cell(row=row, column=2).value = f'Error: Data in Record #{instances} could be displayed incorrect'
                    design(rob, rob.cell(row=row, column=2), start_row=row, end_row=row, start_column=2, end_column=4, constant=3)
                    rob.cell(row=row, column=2).fill = PatternFill(start_color='99FF66', end_color='99FF66', fill_type="solid")
                    row += 2
                rob_string = ''
                row += 3
                instances += 1
                instance_label = rob.cell(row=row - 2, column=2)
                instance_label.value = f'Record #{instances}'
                design(rob=rob, label=instance_label, start_column=2, start_row=row - 2, end_column=11,
                       end_row=row - 2, constant=10, style='thick')
                instance_label.fill = PatternFill(start_color='FF8A09', end_color='FF8A09', fill_type="solid")
                instance_label.font = Font(bold=True)
                continue
        if count == 2 and rob_string in Rob_codes.keys():
            length_limit = int(data[0], base=16) + 3
        put_value = rob.cell(row=row + 1, column=j)
        put_value.value = data[0].strip(' ')
        put_value.alignment = Alignment(horizontal='center')

        # if rob is wakeup/sleep group
        if rob_string in wakeups_rob_code and count > 2:
            # creating values for 'Names of running features'
            if binary_count <= 7:
                if binary_count % 2 == 1:
                    # since even bytes(0,2,4,6) is supported features and odd bytes(1,3,5,7) are running features
                    binary = str(bin(int(put_value.value, base=16))[2:].zfill(len(put_value.value[2:]) * 4))
                    # converting e.g. 0x07 to 0111
                    for index, digits in enumerate(binary):
                        if digits == '1':
                            # since we only record the bits that are '1'
                            merge_count += int(wakeup_fail_etc[binary_count][index][-1])
                            wakeup = wakeup + wakeup_fail_etc[binary_count][index][:-2] + ', '
                    merge_list.append(merge_count)
                    wakeup += '\n'
                    merge_count = 0
            binary_count += 1
            put_value.value = wakeup[:-3]
            if binary_count == 20:
                rob.cell(row=row, column=j - 19).value = 'Names of running features'
                rob.cell(row=row + 1, column=j - 19).value = rob.cell(row=row + 1, column=j).value
                for i in range(1, 20):
                    rob.cell(row=row + 1, column=j - 19 + i).value = ' '
                merge_cell = max(merge_list)
                design(rob=rob, label=rob.cell(row=row + 1, column=j - 19), start_row=row + 1, start_column=j - 19,
                       end_column=j - 19 + merge_cell, end_row=row + 1, constant=merge_cell + 1)
                rob.cell(row=row + 1, column=j - 19).alignment = Alignment(wrap_text=True)
                rob.row_dimensions[row + 1].height = 60

                # creating header for 'Name of running features'
                design(rob=rob, label=rob.cell(row=row, column=j - 19), start_row=row, start_column=j - 19,
                       end_row=row, end_column=j - 19 + merge_cell, constant=merge_cell + 1)
                rob.cell(row=row, column=j - 19).font = Font(bold=True)
        else:
            if rob_string == '0x050x01' and count == 9:
                rob.cell(row=row + 1, column=j - 7).value = int(rob.cell(row=row + 1, column=j - 7).value, base=16)
                rob.cell(row=row + 1, column=j - 6).value = (int(rob.cell(row=row + 1, column=j - 6).value +
                                                                 rob.cell(row=row + 1, column=j - 5).value[2:],
                                                                 base=16))
                rob.merge_cells(start_row=row + 1, end_row=row + 1, start_column=j - 6, end_column=j - 5)
                rob.cell(row=row + 1, column=j - 4).value = int(rob.cell(row=row + 1, column=j - 4).value +
                                                                rob.cell(row=row + 1, column=j - 3).value[2:] +
                                                                rob.cell(row=row + 1, column=j - 2).value[2:] +
                                                                rob.cell(row=row + 1, column=j - 1).value[2:], base=16)
                rob.merge_cells(start_row=row + 1, end_row=row + 1, start_column=j - 4, end_column=j - 1)
            elif rob_string == '0x050x07' and count >= 2:
                put_value.value = int(put_value.value[2], base=16) * 10 + int(put_value.value[3], base=16)
                if put_value.value == 165:
                    put_value.value = 255
            put_value.border = Border(top=Side(border_style="thin"), right=Side(border_style="thin"),
                                      bottom=Side(border_style="thin"), left=Side(border_style="thin"))
        data = data[1:]
        j += 1
        count += 1
        if binary_count == 20:
            binary_count = 0
            merge_list = []
            wakeup = ''
        if count == length_limit and count != 0:
            row += 3
            j = 2
            count = 0
            length_limit = 0
            rob_string = ''
    total = row + 2


# creating a small table that includes the RoB and reason of failure
def column_for_fails(name, rob):
    global total
    column = 2
    total += 2
    rob.cell(row=total, column=1).value = 'NEGATIVE RESPONSES:'
    rob.cell(row=total, column=1).font = Font(bold=True)
    if name:
        for names, fault in name:
            rob.cell(row=total + 1, column=column).value = f'{names}: {fault}'
            design(rob=rob, label=rob.cell(row=total + 1, column=column), start_column=column,
                   start_row=total + 1, end_column=column + 5, end_row=total + 1, constant=6)
            total += 1
    else:
        total += 1
        fail = rob.cell(row=total, column=column)
        fail.value = 'No RoBs with negative responses found'
        design(rob, fail, start_row=total, end_row=total, start_column=column, end_column=column + 3, constant=4)
        fail.font = Font(bold=True)


# sifting out the initial data
def create_excel_and_paste_data(rob, lines, line_number, i, name):
    # filtering out Rx	        EndOfServiceIteration0	    0x00
    #               Rx	            RecordNumbersAndSnapshots	    0x00
    #               Rx	                RecordNumbersAndSnapshots	    0x00
    if i == len(line_number) - 1:
        newlines = lines[line_number[i]:]
    else:
        newlines = lines[line_number[i]:line_number[i + 1]]
    conditions = ['EndOfService', 'RecordNumbers']
    data = []
    for index, line in enumerate(newlines):
        for condition in conditions:
            if condition.lower() in line.lower():
                line_list = line.split('\t')
                data.append(line_list)

    # initial data
    # there will be a column that contains 0x00, 0x00, 0x00, 0x01, 0x01, 0x01
    # the column is located at the extreme right side of the Excel file to prevent the data from being overlapped
    # accidentally when creating the table
    # right now the table is located at column #len of data so the length of each table will never reach that column
    # then after computing table, the column with the data is deleted
    list_of_numbers = []
    for i in range(0, len(data), 3):
        list_of_numbers.append(data[i][2])
    if list_of_numbers:
        list_of_numbers.append(None)
        create_numbers(rob, list_of_numbers, name)


def dtc_design(dtc, label, start_row, start_column, end_row, end_column, constant, style='thin'):
    dtc.merge_cells(start_row=start_row, start_column=start_column, end_row=end_row,
                    end_column=end_column)
    label.alignment = Alignment(horizontal='center')
    for i in range(0, constant):
        dtc.cell(row=start_row, column=start_column + i).border = Border(top=Side(border_style=style),
                                                                         right=Side(border_style=style),
                                                                         bottom=Side(border_style=style),
                                                                         left=Side(border_style=style))


def filtering():
    ############# PHASE 4 STANDARD #############
    for i in reversed(range(len(phase_4_standard_list))):
        if phase_4_standard_list[i][3] == '1':
            phase_4_standard_list[i - 1] += phase_4_standard_list[i][10:]
            phase_4_standard_list.pop(i)
            continue
        j = 0
        if len(phase_4_standard_list[i]) > 24:
            j = 2
        phase_4_standard['ECU'].insert(0, hex(int(phase_4_standard_list[i][6:8], base=16) - 8))
        phase_4_standard['Response'].insert(0, phase_4_standard_list[i][10 + j:12 + j])
        phase_4_standard['Number of DTC'].insert(0, phase_4_standard_list[i][12 + j:14 + j])
        phase_4_standard['DTC'].insert(0, phase_4_standard_list[i][14 + j:])
        while (len(phase_4_standard['DTC'][0])) % 4 != 0 or phase_4_standard['DTC'][0][-4:] == '0000':
            phase_4_standard['DTC'][0] = phase_4_standard['DTC'][0][:-1]
        if len(phase_4_standard['ECU']) != len(set(phase_4_standard['ECU'])):
            for j in range(1,):
                if phase_4_standard['ECU'][j] == phase_4_standard['ECU'][0]:
                    if phase_4_standard['Response'][j] == phase_4_standard['Response'][0] and phase_4_standard['Number of DTC'][j] == phase_4_standard['Number of DTC'][0] and phase_4_standard['DTC'][j] == phase_4_standard['DTC'][0]:
                        phase_4_standard['ECU'].pop(j)
                        phase_4_standard['Response'].pop(j)
                        phase_4_standard['Number of DTC'].pop(j)
                        phase_4_standard['DTC'].pop(j)

    ############# PHASE 4 EXTENDED #############
    for i in reversed(range(len(phase_4_extended_list))):
        if phase_4_extended_list[i][3] == '1':
            phase_4_extended_list[i - 1] += phase_4_extended_list[i][12:]
            phase_4_extended_list.pop(i)
            continue
        j = 0
        if len(phase_4_extended_list[i]) > 24:
            j = 2
        phase_4_extended['ECU'].insert(0, f'0x{phase_4_extended_list[i][8:10]}')
        phase_4_extended['Response'].insert(0, phase_4_extended_list[i][12 + j:14 + j])
        phase_4_extended['Number of DTC'].insert(0, phase_4_extended_list[i][14 + j:16 + j])
        phase_4_extended['DTC'].insert(0, phase_4_extended_list[i][16 + j:])
        while (len(phase_4_extended['DTC'][0])) % 4 != 0 or phase_4_extended['DTC'][0][-4:] == '0000':
            phase_4_extended['DTC'][0] = phase_4_extended['DTC'][0][:-1]
        if len(phase_4_extended['ECU']) != len(set(phase_4_extended['ECU'])):
            for j in range(1,):
                if phase_4_extended['ECU'][j] == phase_4_extended['ECU'][0]:
                    if phase_4_extended['Response'][j] == phase_4_extended['Response'][0] and phase_4_extended['Number of DTC'][j] == phase_4_extended['Number of DTC'][0] and phase_4_extended['DTC'][j] == phase_4_extended['DTC'][0]:
                        phase_4_extended['ECU'].pop(j)
                        phase_4_extended['Response'].pop(j)
                        phase_4_extended['Number of DTC'].pop(j)
                        phase_4_extended['DTC'].pop(j)

    ############# PHASE 5 STANDARD #############
    for i in reversed(range(len(phase_5_standard_list))):
        if phase_5_standard_list[i][3] == '1':
            phase_5_standard_list[i - 1] += phase_5_standard_list[i][10:]
            phase_5_standard_list.pop(i)
            continue
        j = 0
        if len(phase_5_standard_list[i]) > 24:
            j = 2
        phase_5_standard['Response'].insert(0, phase_5_standard_list[i][2:4])
        phase_5_standard['ECU'].insert(0, hex(int(phase_5_standard_list[i][6:8], base=16) - 8))
        phase_5_standard['SID'].insert(0, phase_5_standard_list[i][10 + j:12 + j])
        phase_5_standard['Sub-Function'].insert(0, phase_5_standard_list[i][12 + j:14 + j])
        phase_5_standard['DTCStatusAvailabilityMask'].insert(0, phase_5_standard_list[i][14 + j:16 + j])
        phase_5_standard['DTC'].insert(0, phase_5_standard_list[i][16 + j:])
        while (len(phase_5_standard['DTC'][0])) % 8 != 0 or phase_5_standard['DTC'][0][-8:] == '00000000':
            phase_5_standard['DTC'][0] = phase_5_standard['DTC'][0][:-1]
        phase_5_standard['Number of DTC'].insert(0, len(phase_5_standard['DTC'][0]) // 8)
        if len(phase_5_standard['ECU']) != len(set(phase_5_standard['ECU'])):
            for j in range(1,):
                if phase_5_standard['ECU'][j] == phase_5_standard['ECU'][0]:
                    if phase_5_standard['Response'][j] == phase_5_standard['Response'][0] and phase_5_standard['Number of DTC'][j] == phase_5_standard['Number of DTC'][0] and phase_5_standard['DTC'][j] == phase_5_standard['DTC'][0] and phase_5_standard['SID'][j] == phase_5_standard['SID'][0] and phase_5_standard['Sub-Function'][j] == phase_5_standard['Sub-Function'][0] and phase_5_standard['DTCStatusAvailabilityMask'][j] == phase_5_standard['DTCStatusAvailabilityMask'][0]:
                        phase_5_standard['ECU'].pop(j)
                        phase_5_standard['Response'].pop(j)
                        phase_5_standard['Number of DTC'].pop(j)
                        phase_5_standard['DTC'].pop(j)
                        phase_5_standard['DTCStatusAvailabilityMask'].pop(j)
                        phase_5_standard['Sub-Function'].pop(j)
                        phase_5_standard['SID'].pop(j)

    ############# PHASE 5 EXTENDED #############
    for i in reversed(range(len(phase_5_extended_list))):
        if phase_5_extended_list[i][3] == '1':
            phase_5_extended_list[i - 1] += phase_5_extended_list[i][12:]
            phase_5_extended_list.pop(i)
            continue
        j = 0
        if len(phase_5_extended_list[i]) > 24:
            j = 2
        phase_5_extended['Response'].insert(0, phase_5_extended_list[i][2:4])
        phase_5_extended['ECU'].insert(0, f'0x{phase_5_extended_list[i][8:10]}')
        phase_5_extended['SID'].insert(0, phase_5_extended_list[i][12 + j:14 + j])
        phase_5_extended['Sub-Function'].insert(0, phase_5_extended_list[i][14 + j:16 + j])
        phase_5_extended['DTCStatusAvailabilityMask'].insert(0, phase_5_extended_list[i][16 + j:18 + j])
        phase_5_extended['DTC'].insert(0, phase_5_extended_list[i][18 + j:])
        while (len(phase_5_extended['DTC'][0])) % 8 != 0 or phase_5_extended['DTC'][0][-8:] == '00000000':
            phase_5_extended['DTC'][0] = phase_5_extended['DTC'][0][:-1]
        phase_5_extended['Number of DTC'].insert(0, len(phase_5_extended['DTC'][0]) // 8)
        if len(phase_5_extended['ECU']) != len(set(phase_5_extended['ECU'])):
            for j in range(1,):
                if phase_5_extended['ECU'][j] == phase_5_extended['ECU'][0]:
                    if phase_5_extended['Response'][j] == phase_5_extended['Response'][0] and phase_5_extended['Number of DTC'][j] == phase_5_extended['Number of DTC'][0] and phase_5_extended['DTC'][j] == phase_5_extended['DTC'][0] and phase_5_extended['SID'][j] == phase_5_extended['SID'][0] and phase_5_extended['Sub-Function'][j] == phase_5_extended['Sub-Function'][0] and phase_5_extended['DTCStatusAvailabilityMask'][j] == phase_5_extended['DTCStatusAvailabilityMask'][0]:
                        phase_5_extended['ECU'].pop(j)
                        phase_5_extended['Response'].pop(j)
                        phase_5_extended['Number of DTC'].pop(j)
                        phase_5_extended['DTC'].pop(j)
                        phase_5_extended['DTCStatusAvailabilityMask'].pop(j)
                        phase_5_extended['Sub-Function'].pop(j)
                        phase_5_extended['SID'].pop(j)


def initial_header(name, additional_data):
    title = dtc.cell(row=1, column=1)
    title.value = name
    dtc_design(dtc=dtc, label=title, start_row=1, start_column=1, end_row=1, end_column=4, constant=4)
    title.fill = PatternFill(start_color='000000', end_color='000000', fill_type="solid")
    title.font = Font(color='fefefe', bold=True)

    try:
        time = search('\d+_\d+', name).group().split('_')
        time[0] = f'{time[0][0:2]}-{time[0][2:4]}-{time[0][4:]}'
        time[1] = f'{time[1][0:2]}:{time[1][2:4]}:{time[1][4:]}'
        time = ' '.join(time)
    except (AttributeError, IndexError):
        time = 'Unable to detect'
    occurrence_time_title = dtc.cell(row=2, column=1)
    occurrence_time_title.value = 'Occurrence time'
    occurrence_time_value = dtc.cell(row=2, column=2)
    occurrence_time_value.value = time

    try:
        counter = search('_\d+_\w.dtc', name).group()
        counter = search('\d+', counter).group().strip('_')
    except (AttributeError, IndexError):
        counter = 'Unable to detect'
    counter_title = dtc.cell(row=3, column=1)
    counter_title.value = 'Counter'
    counter_value = dtc.cell(row=3, column=2)
    counter_value.value = counter

    try:
        trigger = search('_.\.dtc', name).group()[1]
    except (AttributeError, IndexError):
        trigger = 'a'
    trigger_title = dtc.cell(row=4, column=1)
    trigger_title.value = 'Trigger identifier'
    trigger_value = dtc.cell(row=4, column=2)
    try:
        trigger_value.value = trigger_dict[trigger.lower()]
    except KeyError:
        trigger_value.value = trigger_dict['a']

    acquisition_time_title = dtc.cell(row=5, column=1)
    acquisition_time_title.value = 'Acquisition time'
    acquisition_time_value = dtc.cell(row=5, column=2)
    if len(additional_data) == 38:
        acquisition_time_value.value = additional_data[:16]
    else:
        acquisition_time_value.value = 'Unable to detect'

    positional_information_title = dtc.cell(row=6, column=1)
    positional_information_title.value = 'Positional information'
    positional_information_value = dtc.cell(row=6, column=2)
    if len(additional_data) == 38:
        positional_information_value.value = additional_data[16:]
    else:
        positional_information_value.value = 'Unable to detect'

    dtc.cell(row=2, column=1).border = Border(top=Side(border_style='thin'), right=Side(border_style='thin'),
                                              bottom=Side(border_style='thin'), left=Side(border_style='thin'))
    dtc.cell(row=3, column=1).border = Border(top=Side(border_style='thin'), right=Side(border_style='thin'),
                                              bottom=Side(border_style='thin'), left=Side(border_style='thin'))
    dtc.cell(row=4, column=1).border = Border(top=Side(border_style='thin'), right=Side(border_style='thin'),
                                              bottom=Side(border_style='thin'), left=Side(border_style='thin'))
    dtc.cell(row=5, column=1).border = Border(top=Side(border_style='thin'), right=Side(border_style='thin'),
                                              bottom=Side(border_style='thin'), left=Side(border_style='thin'))
    dtc.cell(row=6, column=1).border = Border(top=Side(border_style='thin'), right=Side(border_style='thin'),
                                              bottom=Side(border_style='thin'), left=Side(border_style='thin'))
    for i in range(0, 5):
        dtc.merge_cells(start_row=2 + i, end_row=2 + i, start_column=2, end_column=4)
        for j in range(0, 3):
            dtc.cell(row=2 + i, column=2 + j).border = Border(top=Side(border_style='thin'),
                                                              right=Side(border_style='thin'),
                                                              bottom=Side(border_style='thin'),
                                                              left=Side(border_style='thin'))


def the_rest(no_group):
    dtc.cell(row=2, column=6).value = f'{len(no_group)} data does not belong anywhere: '
    for data in no_group:
        dtc.cell(row=2, column=6).value += f'{data} '


def phase_4_design(name, phase):
    global total
    # PHASE 4 title
    row = 8
    title = dtc.cell(row=row, column=total)
    title.value = name
    dtc_design(dtc=dtc, label=title, start_row=row, start_column=total, end_column=total + 1, end_row=row, constant=2)
    title.fill = PatternFill(start_color='C65911', end_color='C65911', fill_type="solid")
    title.font = Font(color='fefefe', bold=True)

    if len(phase['ECU']) != 0:
        instance = 1
        row += 2
        column = total
        for x in range(len(phase['ECU'])):
            instance_label = dtc.cell(row=row, column=total)
            instance_label.value = f'Record #{instance}'
            instance_label.fill = PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type="solid")
            dtc_design(dtc=dtc, label=instance_label, start_row=row, start_column=total, end_column=total + 1,
                       end_row=row,
                       constant=2)
            instance_label.font = Font(bold=True)
            instance += 1
            row += 1
            for key in phase.keys():
                if key == 'DTC':
                    if int(phase['Number of DTC'][x], base=16) > 0:
                        header = dtc.cell(row=row, column=column)
                        header.value = key
                        header.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type="solid")
                        dtc.merge_cells(start_row=row, start_column=column,
                                        end_row=row + int(phase['Number of DTC'][x], base=16) - 1,
                                        end_column=column)
                        header.alignment = Alignment(horizontal='center', vertical='center')
                        for i in range(0, int(phase['Number of DTC'][x], base=16)):
                            dtc.cell(row=row + i, column=column).border = Border(top=Side(border_style='thin'),
                                                                                 right=Side(border_style='thin'),
                                                                                 bottom=Side(border_style='thin'),
                                                                                 left=Side(border_style='thin'))

                        for k in range(len(phase['DTC'][x]) // 4):
                            put_value = dtc.cell(row=row, column=column + 1)
                            put_value.value = phase[key][x][4 * k:4 * k + 4]
                            dtc_design(dtc=dtc, label=put_value, start_row=row, start_column=column + 1,
                                       end_column=column + 1,
                                       end_row=row, constant=1)
                            row += 1
                else:
                    header = dtc.cell(row=row, column=column)
                    header.value = key
                    dtc_design(dtc, label=header, start_row=row, start_column=column, end_column=column, end_row=row,
                               constant=2)
                    put_value = dtc.cell(row=row, column=column + 1)
                    put_value.value = phase[key][x]
                    dtc_design(dtc, put_value, start_row=row, start_column=column + 1, end_row=row,
                               end_column=column + 1,
                               constant=1)

                row += 1
                column = total
            row += 1
    else:
        empty = dtc.cell(column=total, row=row + 1)
        empty.value = f'No DTCs found in {name}'
        dtc_design(dtc=dtc, label=empty, start_row=row + 1, start_column=total, end_column=total + 1, end_row=row + 1,
                   constant=2)


def phase_5_design(name, phase):
    global total
    # PHASE 5 title
    row = 8
    title = dtc.cell(row=row, column=total)
    title.value = name
    dtc_design(dtc=dtc, label=title, start_row=row, start_column=total, end_column=total + 3, end_row=row, constant=4)
    title.fill = PatternFill(start_color='305496', end_color='305496', fill_type="solid")
    title.font = Font(color='fefefe', bold=True)

    # RECORD NUMBER
    if len(phase['ECU']) != 0:
        instance = 1
        row += 2
        count = 0
        for i in range(len(phase['ECU'])):
            instance_label = dtc.cell(row=row, column=total)
            instance_label.value = f'Record #{instance}'
            instance_label.fill = PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type="solid")
            dtc_design(dtc=dtc, label=instance_label, start_row=row, start_column=total, end_column=total + 3,
                       end_row=row,
                       constant=4)
            instance_label.font = Font(bold=True)
            instance += 1
            row += 1
            column = total
            for key in phase.keys():
                if key == 'DTC':
                    if phase['Number of DTC'][i] != 0:
                        header = dtc.cell(row=row, column=column)
                        header.value = key
                        dtc_design(dtc=dtc, label=header, start_column=column, start_row=row, end_column=column + 1,
                                   end_row=row,
                                   constant=2)
                        header.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type="solid")
                        status_header = dtc.cell(row=row, column=column + 2)
                        status_header.value = 'Status'
                        dtc_design(dtc=dtc, label=status_header, start_column=column + 2, start_row=row,
                                   end_column=column + 3,
                                   end_row=row,
                                   constant=2)
                        status_header.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type="solid")
                        row += 1
                        for k in range(phase['Number of DTC'][i]):
                            put_value = dtc.cell(row=row, column=column)
                            put_value.value = '0x' + phase['DTC'][i][8 * k: 8 * (k + 1) - 2]
                            dtc_design(dtc=dtc, label=put_value, start_column=column, start_row=row,
                                       end_column=column + 1,
                                       end_row=row, constant=2)
                            put_value = dtc.cell(row=row, column=column + 2)
                            put_value.value = phase['DTC'][i][8 * k + 6: 8 * (k + 1)]
                            dtc_design(dtc=dtc, label=put_value, start_column=column + 2, start_row=row,
                                       end_column=column + 3,
                                       end_row=row, constant=2)
                            row += 1
                else:
                    header = dtc.cell(row=row, column=column)
                    header.value = key
                    dtc_design(dtc=dtc, label=header, start_column=column, start_row=row, end_column=column,
                               end_row=row,
                               constant=1)
                    column += 1
                    put_value = dtc.cell(row=row, column=column)
                    put_value.value = phase[key][i]
                    if key == 'Response':
                        try:
                            put_value.value = Responses[phase[key][i]]
                        except KeyError:
                            put_value.value = Responses['0']
                    dtc_design(dtc=dtc, label=put_value, start_column=column, start_row=row, end_column=column,
                               end_row=row,
                               constant=1)
                    if count % 2 == 0:
                        column += 1
                    else:
                        row += 1
                        column = total
                    count += 1
            row += 2
    else:
        empty = dtc.cell(column=total, row=row + 1)
        empty.value = f'No DTCs found in {name}'
        dtc_design(dtc=dtc, label=empty, start_row=row + 1, start_column=total, end_column=total + 3, end_row=row + 1,
                   constant=4)


def no_dtc():
    global filename
    column = 1
    row = 1
    title = xlsx.cell(row=row, column=column)
    title.value = 'DTC'
    title.font = Font(bold=True)
    design(xlsx, title, start_row=row, end_row=row, start_column=column, end_column=column + 1, constant=2,
           style='thick')
    row += 1
    fail = xlsx.cell(row=row, column=column + 1)
    fail.value = f'No DTCs found in {filename}'
    fail.font = Font(bold=True)


# window pop up will always appear after creating an Excel file until user clicks close
text_errors = ['Select a Text file or \nan Excel file!']
address_errors = ['Select a Destination folder!', 'Select another Destination folder!']
passed = 1
# passed is when user clicks on convert with no errors
while passed:
    date_time_str = datetime.now().strftime("%d%b%y-%H%M%S")
    open_address = ''
    save_address = ''
    filename = ''
    total = 1
    passed = 0
    choice = 0
    xlsx = openpyxl.Workbook()
    window_explorer()
    # choice 1 == RoB, choice 2 == remote DTC
    if choice == 1:
        dtcs, positive_name, negative_name, lines, positive_occurrences = text_file()
        rob = xlsx.active
        dtc_in_rob_design(rob, dtcs)
        total += 2
        rob.cell(row=total, column=1).value = 'RoB'
        design(rob, rob.cell(row=total, column=1), start_row=total, start_column=1, end_column=2, end_row=total,
               constant=2,
               style='thick')
        rob.cell(row=total, column=1).font = Font(bold=True)
        total += 1
        rob.cell(row=total, column=1).value = 'POSITIVE RESPONSES:'
        rob.cell(row=total, column=1).font = Font(bold=True)
        if positive_name:
            for i in range(0, len(positive_name)):
                name_title = rob.cell(row=2 + total, column=1)
                name_title.font = Font(bold=True)
                name_title.value = positive_name[i].upper()
                total += 2
                create_excel_and_paste_data(rob, lines, positive_occurrences, i, positive_name[i].upper())
        else:
            total += 1
            fail = rob.cell(row=total, column=2)
            fail.value = 'No RoBs with positive responses found'
            design(rob, fail, start_row=total, end_row=total, start_column=2, end_column=5, constant=4)
            fail.font = Font(bold=True)
            total += 1
        column_for_fails(negative_name, rob)

        # adjusting width of columns to fit the words
        dim_holder = DimensionHolder(worksheet=rob)
        for col in range(rob.min_column, rob.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(rob, min=col, max=col, width=15)
        rob.column_dimensions = dim_holder
        xlsx.save(f'{save_address}/RoB_DTC-({filename})-{date_time_str}.xlsx')
        os.startfile(f'{save_address}/RoB_DTC-({filename})-{date_time_str}.xlsx')
    elif choice == 2:
        names = {}
        xlsx_dtc = openpyxl.Workbook()
        input_excel = openpyxl.load_workbook(filename=open_address, read_only=True)
        input_sheet = input_excel.active
        pattern = '0c......................'
        additional_data = {}
        xlsx = xlsx_dtc.active
        for i in range(1, input_sheet.max_column+1):
            if input_sheet.cell(row=1, column=i).value is not None and 'filename' in input_sheet.cell(row=1, column=i).value.strip(' ').lower():
                for j in range(1, input_sheet.max_row+1):
                    if input_sheet.cell(row=j, column=i).value is not None and '.dtc' in input_sheet.cell(row=j,
                                                                                                          column=i).value:
                        try:
                            additional_data[input_sheet.cell(row=j, column=i).value] = input_sheet.cell(row=j,
                                                                                                        column=i + 1).value[0:38]
                        except (TypeError, IndexError):
                            additional_data[input_sheet.cell(row=j, column=i).value] = ''
                        try:
                            names[input_sheet.cell(row=j, column=i).value] = findall(pattern,
                                                                                     input_sheet.cell(row=j, column=i + 1).value)
                        except (TypeError, IndexError):
                            names[input_sheet.cell(row=j, column=i).value] = []
                break
        input_excel.close()
        for data in names.keys():
            total = 1
            phase_4_standard_list = []
            phase_4_standard = {'ECU': [], 'Response': [], 'Number of DTC': [], 'DTC': []}
            phase_4_extended_list = []
            phase_4_extended = {'ECU': [], 'Response': [], 'Number of DTC': [], 'DTC': []}
            phase_5_standard_list = []
            phase_5_standard = {'ECU': [], 'Response': [], 'SID': [],
                                'Sub-Function': [], 'Number of DTC': [], 'DTCStatusAvailabilityMask': [], 'DTC': []}
            phase_5_extended_list = []
            phase_5_extended = {'ECU': [], 'Response': [], 'SID': [],
                                'Sub-Function': [], 'Number of DTC': [], 'DTCStatusAvailabilityMask': [], 'DTC': []}
            no_group = []
            for DTC in names[data]:
                if '1' <= DTC[2] <= '5':
                    phase_4_standard_list.append(DTC)
                elif '6' <= DTC[2] <= '8':
                    phase_4_extended_list.append(DTC)
                elif '9' == DTC[2] or 'a' <= DTC[2].lower() <= 'c':
                    phase_5_standard_list.append(DTC)
                elif 'd' <= DTC[2].lower() <= 'f':
                    phase_5_extended_list.append(DTC)
                else:
                    no_group.append(DTC)
            filtering()
            dtc = xlsx_dtc.create_sheet(data)
            initial_header(data, additional_data[data])
            phase_5_design('Phase 5 (standard)', phase_5_standard)
            total += 5
            phase_5_design('Phase 5 (extended)', phase_5_extended)
            total += 5
            phase_4_design('Phase 4 (extended)', phase_4_extended)
            total += 3
            phase_4_design('Phase 4 (standard)', phase_4_standard)
            if no_group:
                the_rest(no_group)
            dim_holder = DimensionHolder(worksheet=dtc)
            for col in range(dtc.min_column, dtc.max_column + 1):
                dim_holder[get_column_letter(col)] = ColumnDimension(dtc, min=col, max=col, width=25)
            dtc.column_dimensions = dim_holder
        if names:
            xlsx_dtc.remove(xlsx_dtc['Sheet'])
        else:
            no_dtc()
        xlsx_dtc.save(f'{save_address}/Remote_DTC-({filename})-{date_time_str}.xlsx')
        os.startfile(f'{save_address}/Remote_DTC-({filename})-{date_time_str}.xlsx')
